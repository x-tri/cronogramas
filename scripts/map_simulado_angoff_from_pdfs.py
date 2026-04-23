#!/usr/bin/env python3
"""Gera planilha XLSX de mapeamento de simulado a partir de PDFs reais.

Fonte dos dados:
- caderno LC/CH
- gabarito LC/CH
- caderno CN/MT
- gabarito CN/MT

Saida:
- planilha Excel com as colunas solicitadas pelo usuario
- aba adicional de metadados para auditoria

Observacao importante:
- a coluna "Dificuldade ANGOFF" e uma estimativa tecnica ordinal (1..5)
  baseada no enunciado real e no assunto do item.
- ela nao e um campo oficial do INEP.
"""

from __future__ import annotations

import argparse
import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from textwrap import dedent

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pypdf import PdfReader


ROOT = Path("/Users/home/Desktop/MARISTA/1o SIMULADO DA MENTORIA")
DEFAULT_LC_CH_CADERNO = ROOT / "LC e CH - Caderno.pdf"
DEFAULT_LC_CH_GABARITO = ROOT / "LC e CH - gabarito.pdf"
DEFAULT_CN_MT_CADERNO = ROOT / "CN e MAT - Caderno_.pdf"
DEFAULT_CN_MT_GABARITO = ROOT / "CN e MAT - gabarito.pdf"
DEFAULT_OUTPUT = ROOT / "mapeamento-angoff-simulado-2-2025.xlsx"


QUESTION_MARKER = re.compile(r"QUEST[ÃA]O\s+(\d{2,3})", flags=re.IGNORECASE)
ANSWER_MARKER = re.compile(r"(\d{1,3})\s*[\u200a\u2009\u2008\u2007\u202f\u00a0 ]*([ABCDE])")


@dataclass(frozen=True)
class SimuladoPaths:
    lc_ch_caderno: Path
    lc_ch_gabarito: Path
    cn_mt_caderno: Path
    cn_mt_gabarito: Path


TOPIC_ROWS = dedent(
    """
    # LINGUAGENS (1-45)
    1|Ingles - anuncio sobre consumo sustentavel de oleo de palma|2
    2|Ingles - poema sobre resistencia negra e legado da escravidao|2
    3|Ingles - poema sobre exemplo pratico e persuasao moral|2
    4|Ingles - poema sobre preconceito contra latinos e identidade|3
    5|Ingles - linguagem verbal e visual em dialogo humoristico|2
    6|Literatura - reiteracao e progressao textual em cancao|2
    7|Literatura - ironia e violencia escravocrata em Machado de Assis|3
    8|Leitura de midia - fenomeno dos bebes reborn e repercussao social|2
    9|Literatura africana - musica como resistencia e liberdade|3
    10|Linguagens - campanha de conscientizacao sobre autismo|1
    11|Literatura contemporanea - poesia feminista e critica a violencia de genero|3
    12|Esporte e sociedade - representatividade feminina no esporte|2
    13|Literatura afro-brasileira - ancestralidade e oralidade em narrativa|3
    14|Linguagem digital - coloquialidade e viralizacao de Machado no TikTok|2
    15|Artes - arte africana e afro-brasileira, ancestralidade e coletividade|2
    16|Literatura contemporanea - fuga feminina e subjetividade|3
    17|Linguagens indigenas - lingua materna e identidade dos povos originarios|2
    18|Literatura afro-brasileira - resistencia feminina e ancestralidade em narrativa|4
    19|Cultura popular - maracatu e valorizacao das tradicoes brasileiras|2
    20|Poema contemporaneo - autoimagem e conflito identitario|2
    21|Publicidade - critica a interferencia da industria de ultraprocessados|3
    22|Midia e esporte - racismo no futebol e discurso antirracista|2
    23|Literatura e ditadura - linguagem infantil para critica social e politica|5
    24|Artes visuais - Frida Kahlo, identidade e resistencia|3
    25|Poesia - reconhecimento do eu na relacao com o outro|3
    26|Filosofia da linguagem - vaidade e apego a opiniao em aforismo|5
    27|Literatura naturalista - racismo e preconceito na sociedade brasileira|3
    28|Genero jornalistico - premiacao cinematografica e objetividade da noticia|2
    29|Modernismo - critica a burguesia e ruptura estetica|3
    30|Variacao linguistica - bilinguismo hispanico e identidade cultural|3
    31|Poesia modernista - indiferenca social diante da guerra|4
    32|Metalinguagem - uso e crise do ponto e virgula|2
    33|Guimaraes Rosa - construcao de personagem e linguagem literaria|5
    34|Texto expositivo - impactos da inteligencia artificial na industria|2
    35|Musica brasileira - sincretismo cultural e matrizes musicais|2
    36|Narrativa fantastica - humor e quebra de expectativa|3
    37|Publicidade - campanha contra violencia domestica e feminicidio|2
    38|Cancao - dialogo entre cultura afro-brasileira e linguagem digital|3
    39|Cultura digital - hashtags e organizacao da informacao|2
    40|Literatura de cordel - invencao, sonho e tradicao popular|3
    41|Arte urbana - grafite e xilogravura nordestina|3
    42|Corpo e movimento - consciencia corporal e percepcao dos gestos|3
    43|Letra de samba - aconselhamento amoroso e funcao conativa|2
    44|Campanha de saude - funcao referencial em peca sobre vacina|1
    45|Sociologia - racismo como elemento cultural|3

    # HUMANAS (46-90)
    46|Brasil colonial - confrarias negras e organizacao comunitaria|2
    47|Brasil colonial - poder dos grandes proprietarios e hierarquia social|2
    48|Geografia agraria - agronegocio no Cerrado e subordinacao do campesinato|3
    49|Geografia agraria - defensivos agricolas e impactos socioambientais|2
    50|Brasil colonial - controle social e marginalizacao dos ciganos|3
    51|Cultura - diversidade cultural e organizacao socioespacial|2
    52|Segundo Reinado - ciencia, tecnologia e imagem de modernidade do Imperio|2
    53|Sociologia marxista - ideologia e meios de comunicacao|3
    54|Geopolitica - disputa territorial na Caxemira|3
    55|Historia contemporanea - legado social da Revolucao Francesa|3
    56|Geografia economica - Industria 4.0 e transformacoes do trabalho|3
    57|Cultura popular - memoria coletiva e transmissao de saberes|2
    58|Filosofia antiga - relativismo sofista e universalismo socratico|3
    59|Historia mundial - acordos imperialistas e Oriente Medio na Primeira Guerra|5
    60|Ditadura militar - repressao aos movimentos sociais e censura|2
    61|Geomorfologia - rios como agentes externos do relevo|2
    62|Ditadura militar - censura cultural e controle ideologico|2
    63|Migracoes - midia, refugiados e xenofobia|2
    64|Geopolitica - rivalidade economica entre EUA e China|3
    65|Etica e democracia - responsabilidade coletiva dos cidadaos|2
    66|Urbanizacao - desigualdade socioespacial nas cidades brasileiras|2
    67|Movimentos sociais - mulheres negras, racismo e machismo estrutural|3
    68|Guerra Fria - conflitos por procuracao no Terceiro Mundo|3
    69|Filosofia social - Foucault, vigilancia e controle institucional|5
    70|Idade Media - escolastica, universidade e humanismo cristao|4
    71|Meio ambiente - Acordo de Paris e transicao para baixo carbono|2
    72|America pre-colombiana - Imperio Inca e colonizacao espanhola|3
    73|Historia ambiental - privatizacao das florestas e expansao maritima|3
    74|Energia - fontes alternativas e diversificacao da matriz energetica|2
    75|Regioes do Brasil - desigualdades regionais e ocupacao do territorio|2
    76|Filosofia feminista - justica e desigualdade de genero|5
    77|Cartografia - escala e nivel de detalhamento do mapa|2
    78|Brasil colonial - patriarcalismo luso-brasileiro e poder social|3
    79|Historia da ciencia no Brasil - vacinacao e iniciativa de agentes locais|4
    80|Movimentos negros - combate ao racismo estrutural|2
    81|Filosofia politica - contrato social e vontade geral em Rousseau|3
    82|Republica Velha - coronelismo e voto de cabresto|2
    83|Geografia economica - blocos economicos e integracao regional|2
    84|Era Vargas - crise politica de 1954 e suicidio de Vargas|3
    85|Urbanizacao brasileira - segregacao e deficit de infraestrutura|2
    86|Empirismo - Locke, experiencia e formacao cultural|3
    87|Globalizacao - comercio internacional e organizacao do espaco|2
    88|Cidadania - desigualdade estrutural e acao coletiva|2
    89|Racionalismo - duvida metodica e cogito cartesiano|3
    90|Interpretacao do Brasil - mestiagem e formacao cultural brasileira|3

    # NATUREZA (91-135)
    91|Fisica - velocidade media com tempos iguais|2
    92|Quimica - estequiometria da sintese de ETBE a partir do etanol|5
    93|Biologia - resistencia a insulina e producao de androgenos na SOP|3
    94|Fisica - lancamento vertical e aceleracao da gravidade|2
    95|Quimica - equilibrio acido-base e indicador antocianina|5
    96|Biologia - sodio, impulso nervoso e equilibrio hidrico|2
    97|Fisica - velocidade media entre sensores de radar|2
    98|Quimica - hidrolise de tensoativos e impacto ambiental|4
    99|Biologia vegetal - fotossintese C4 e adaptacao ao clima quente|3
    100|Fisica - tracao e equilibrio de forcas no cabo de guerra|3
    101|Quimica organica - quiralidade e carbono assimetrico|3
    102|Biotecnologia - vacinas de mRNA e inducao de imunidade|2
    103|Fisica - leis de Newton e seguranca em colisoes|3
    104|Quimica organica - polaridade e solubilidade da caseina|3
    105|Biologia vegetal - sementes e evolucao das fanerogamas|3
    106|Fisica - associacao em paralelo de lampadas e corrente eletrica|2
    107|Biologia vegetal - xilema e extracao de celulose|2
    108|Quimica ambiental - solubilidade de pesticidas e mobilidade no ambiente|3
    109|Fisica - inducao eletromagnetica em carregamento por inducao|3
    110|Quimica nuclear - radioatividade do potassio-40 e equivalencia de dose|3
    111|Biologia celular - micronucleos e microscopia de fluorescencia|3
    112|Fisica ondulatoria - interferencia de ondas sonoras|3
    113|Eletroquimica - galvanoplastia e leis de Faraday|5
    114|Ecologia - bioacumulacao de metais pesados na cadeia alimentar|2
    115|Fisica ondulatoria - batimentos sonoros e frequencia|3
    116|Quimica organica - cromoforo e conjugacao eletronica|4
    117|Ecologia - queimadas e perda da materia organica do solo|3
    118|Fisica ondulatoria - refracao das ondas do mar|3
    119|Fisica termica - condensacao do vapor de agua em extintor de CO2|3
    120|Ecologia - desmatamento tropical e liberacao de carbono|2
    121|Fisica moderna - radiacao ultravioleta e ionizacao|2
    122|Ecologia - corredores ecologicos e conservacao do tamandua-bandeira|2
    123|Quimica organica - sintese do tamoxifeno e reacao organica|4
    124|Fisica das radiacoes - exposicao ocupacional a radiacao cosmica em voos|2
    125|Quimica - ligacao quimica e minimo de energia potencial em H2|3
    126|Biologia - fertilizacao in vitro e reproducao assistida|2
    127|Fisico-quimica - equacao dos gases ideais e armazenamento de metano|3
    128|Quimica - solucao saturada, densidade e extracao de microplasticos|5
    129|Histologia - tecido muscular estriado esqueletico e movimento voluntario|2
    130|Quimica e fisica - densidade de solucoes e identificacao de metal|4
    131|Fisica - compressao isotermica e lei de Boyle|2
    132|Biologia celular - respiracao aerobica e sintese de ATP na mitocondria|3
    133|Fisica termica - calor sensivel e potencia de aquecedor|3
    134|Biologia - prevencao da leishmaniose visceral|2
    135|Fisica moderna - fissao nuclear do uranio e geracao de energia|3

    # MATEMATICA (136-180)
    136|Matematica - conversao de unidades de comprimento e divisao|1
    137|Matematica - calculo de tempo total com intervalos|2
    138|Matematica - combinacao de alimentos e reducao proporcional de carboidratos|3
    139|Matematica - razao e proporcionalidade no consumo de tintas|3
    140|Matematica - notacao cientifica e conversao para nanometros|2
    141|Matematica - volume com cavidade e massa de paralelepipedo|2
    142|Matematica - sistema de equacoes com lucro e prejuizo|3
    143|Matematica - media, mediana e moda em tabela de frequencias|2
    144|Matematica - leitura de numero romano|1
    145|Matematica - analise combinatoria na formacao de juri|3
    146|Matematica - porcentagem aplicada a doencas pulmonares|2
    147|Matematica - escala e proporcionalidade em miniaturas|2
    148|Matematica - comparacao de areas para reserva ambiental minima|2
    149|Matematica - funcao por partes e variacao linear de preco|5
    150|Matematica - probabilidade total para chuva no sabado|3
    151|Matematica - parabola e leitura de valor em grafico|3
    152|Matematica - projecao ortogonal de objeto tridimensional|4
    153|Matematica - comparacao de consumo e custo de combustiveis|3
    154|Matematica - proporcionalidade em modelo de custo energetico|3
    155|Matematica - desconto percentual em plano de saude familiar|2
    156|Matematica - gasto calorico e conversao para perda de massa|3
    157|Matematica - extrapolacao linear em grafico de internacoes|3
    158|Matematica - volume de cilindro com semiesfera e resto de solucao|3
    159|Matematica - lei de Benford e logaritmo decimal|5
    160|Matematica - desvio padrao e regularidade estatistica|1
    161|Matematica - identificacao de solidos geometricos em objetos|1
    162|Matematica - expressao algebrica do valor total com desconto e frete|1
    163|Matematica - taxa de desemprego a partir de PEA e PO|2
    164|Matematica - area superficial com escala em gazebo|5
    165|Matematica - permutacao com repeticao de campeoes|5
    166|Matematica - media ponderada e variacao cambial|3
    167|Matematica - funcao do faturamento e valor maximo|4
    168|Matematica - regra de tres composta com caminhoes e tempo|3
    169|Matematica - classificacao de numeros em conjuntos numericos|3
    170|Matematica - consumo de energia em kWh e comparacao diaria|3
    171|Matematica financeira - quitacao antecipada e desconto implicito|5
    172|Matematica - leitura de grafico para escolha do menor preco|2
    173|Matematica - media aritmetica com termo faltante|2
    174|Matematica - distribuicao binomial para exatamente quatro falhas|4
    175|Matematica - comprimento de arco e comparacao de configuracoes|2
    176|Matematica - logaritmo e entropia minima de senha|5
    177|Matematica - funcao lucro com custo fixo e variavel|2
    178|Matematica - funcao seno e mes de menor preco|4
    179|Matematica - padrao de quadrados e sequencia numerica|3
    180|Matematica - relacao entre volumes de cone e cilindro|2
    """
).strip()


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Gera planilha Excel com mapeamento de conteudo e dificuldade Angoff.",
    )
    parser.add_argument("--lc-ch-caderno", type=Path, default=DEFAULT_LC_CH_CADERNO)
    parser.add_argument("--lc-ch-gabarito", type=Path, default=DEFAULT_LC_CH_GABARITO)
    parser.add_argument("--cn-mt-caderno", type=Path, default=DEFAULT_CN_MT_CADERNO)
    parser.add_argument("--cn-mt-gabarito", type=Path, default=DEFAULT_CN_MT_GABARITO)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    return parser.parse_args()


def extract_text(pdf_path: Path) -> str:
    reader = PdfReader(str(pdf_path))
    return "\n".join((page.extract_text() or "") for page in reader.pages)


def parse_question_markers(pdf_path: Path) -> list[int]:
    text = extract_text(pdf_path)
    return [int(n) for n in QUESTION_MARKER.findall(text)]


def parse_answer_key(pdf_path: Path) -> dict[int, str]:
    text = extract_text(pdf_path)
    pairs = ANSWER_MARKER.findall(text)
    answers: dict[int, str] = {}
    for raw_number, letter in pairs:
        number = int(raw_number)
        if 1 <= number <= 180:
            answers[number] = letter
    return answers


def parse_topic_rows(raw: str) -> dict[int, tuple[str, int]]:
    rows: dict[int, tuple[str, int]] = {}
    for line in raw.splitlines():
        item = line.strip()
        if not item or item.startswith("#"):
            continue
        try:
            raw_number, topic, raw_difficulty = item.split("|", 2)
        except ValueError as exc:
            raise ValueError(f"Linha invalida no mapeamento: {item}") from exc
        number = int(raw_number)
        difficulty = int(raw_difficulty)
        rows[number] = (topic.strip(), difficulty)
    return rows


def validate_sources(paths: SimuladoPaths) -> None:
    for path in (
        paths.lc_ch_caderno,
        paths.lc_ch_gabarito,
        paths.cn_mt_caderno,
        paths.cn_mt_gabarito,
    ):
        if not path.exists():
            raise FileNotFoundError(f"Arquivo nao encontrado: {path}")

    lc_questions = parse_question_markers(paths.lc_ch_caderno)
    cn_questions = parse_question_markers(paths.cn_mt_caderno)
    if lc_questions != list(range(1, 91)):
        raise RuntimeError(
            f"Caderno LC/CH com marcacao inesperada. Obtido: {lc_questions[:5]}...{lc_questions[-5:]}",
        )
    if cn_questions != list(range(91, 181)):
        raise RuntimeError(
            f"Caderno CN/MT com marcacao inesperada. Obtido: {cn_questions[:5]}...{cn_questions[-5:]}",
        )

    answers: dict[int, str] = {}
    answers.update(parse_answer_key(paths.lc_ch_gabarito))
    answers.update(parse_answer_key(paths.cn_mt_gabarito))
    if sorted(answers) != list(range(1, 181)):
        raise RuntimeError("Gabarito oficial incompleto. Era esperado um total de 180 itens.")


def build_rows(paths: SimuladoPaths) -> list[dict[str, object]]:
    answers: dict[int, str] = {}
    answers.update(parse_answer_key(paths.lc_ch_gabarito))
    answers.update(parse_answer_key(paths.cn_mt_gabarito))

    topic_map = parse_topic_rows(TOPIC_ROWS)
    if sorted(topic_map) != list(range(1, 181)):
        raise RuntimeError("Mapa de conteudo/dificuldade incompleto. Era esperado um total de 180 itens.")

    rows: list[dict[str, object]] = []
    for number in range(1, 181):
        topic, difficulty = topic_map[number]
        if difficulty < 1 or difficulty > 5:
            raise RuntimeError(f"Dificuldade fora da escala 1..5 na questao {number}.")
        rows.append(
            {
                "Numero da questao (1 a 180)": number,
                "Conteúdo ENEM resumido": topic,
                "Gabarito oficial (letra)": answers[number],
                "Dificuldade ANGOFF": difficulty,
            },
        )
    return rows


def autosize_columns(ws) -> None:
    for column_cells in ws.columns:
        max_len = 0
        column_index = column_cells[0].column
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        width = min(max_len + 2, 72)
        ws.column_dimensions[get_column_letter(column_index)].width = width


def style_header(ws, row: int = 1) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[row]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def write_workbook(paths: SimuladoPaths, rows: list[dict[str, object]], output_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "itens"

    headers = [
        "Numero da questao (1 a 180)",
        "Conteúdo ENEM resumido",
        "Gabarito oficial (letra)",
        "Dificuldade ANGOFF",
    ]
    ws.append(headers)
    for row in rows:
        ws.append([row[h] for h in headers])

    style_header(ws)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:D{ws.max_row}"
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 72
    ws.column_dimensions["C"].width = 24
    ws.column_dimensions["D"].width = 20
    for data_row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        data_row[0].alignment = Alignment(horizontal="center")
        data_row[1].alignment = Alignment(wrap_text=True, vertical="top")
        data_row[2].alignment = Alignment(horizontal="center")
        data_row[3].alignment = Alignment(horizontal="center")

    meta = wb.create_sheet("metadados")
    meta_rows = [
        ("Fonte dos dados", "PDFs reais fornecidos pelo usuario"),
        ("N amostral", "180 questoes"),
        ("Escala Angoff", "1=fácil, 2=fácil-medio, 3=medio, 4=medio-dificil, 5=dificil"),
        (
            "Observacao",
            "A dificuldade ANGOFF foi estimada tecnicamente a partir do enunciado real e do assunto do item; nao e um campo oficial do INEP.",
        ),
        ("Arquivo LC/CH - Caderno", str(paths.lc_ch_caderno)),
        ("Arquivo LC/CH - Gabarito", str(paths.lc_ch_gabarito)),
        ("Arquivo CN/MT - Caderno", str(paths.cn_mt_caderno)),
        ("Arquivo CN/MT - Gabarito", str(paths.cn_mt_gabarito)),
        ("Gerado em", datetime.now().isoformat(timespec="seconds")),
    ]
    meta.append(["Campo", "Valor"])
    for item in meta_rows:
        meta.append(list(item))
    style_header(meta)
    meta.freeze_panes = "A2"
    meta.column_dimensions["A"].width = 28
    meta.column_dimensions["B"].width = 120
    for row in meta.iter_rows(min_row=2, max_row=meta.max_row):
        row[0].font = Font(bold=True)
        row[1].alignment = Alignment(wrap_text=True, vertical="top")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main() -> int:
    args = parse_args()
    paths = SimuladoPaths(
        lc_ch_caderno=args.lc_ch_caderno,
        lc_ch_gabarito=args.lc_ch_gabarito,
        cn_mt_caderno=args.cn_mt_caderno,
        cn_mt_gabarito=args.cn_mt_gabarito,
    )

    validate_sources(paths)
    rows = build_rows(paths)
    write_workbook(paths, rows, args.output)

    print(f"Planilha gerada com sucesso: {args.output}")
    print("Fonte dos dados: 4 PDFs reais")
    print(f"N amostral: {len(rows)} questoes")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
